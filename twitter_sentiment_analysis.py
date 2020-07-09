##get all tweets between two dates from a specified twitter user
#we have a limit of 3200 tweets for twitter api after that we get an error
# we have option of choosing timeline as startdate and endDate
import openpyxl
import sys,tweepy,csv,re
from textblob import TextBlob
import matplotlib.pyplot as plt
import datetime
import xlsxwriter
import sys
import tweepy
from tweepy import Stream
from tweepy import OAuthHandler
from tweepy.streaming import StreamListener
import json
import pandas as pd
import csv
import re #regular expression
from textblob import TextBlob
import string
import preprocessor as p
from textblob import TextBlob
from googletrans import Translator
import emoji
import nltk
from mtranslate import translate
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from wordcloud import WordCloud, STOPWORDS
from matplotlib.dates import date2num
import matplotlib.dates as mdates

class SentimentAnalysis():

    def __init__(self):
        self.tweets = []
        self.tweetText = []
        self.final_tweets=[]
    
    #add multiple strings as we have a constraint of 250 api calls for google translate for free
    def translator_text(self, final_tweets, ids, texts, lang):
        text_trans=""
        all_text=""
        trans=[]
        print(len(texts))
        #convert to one string
        for i in range(0, len(texts)):
            print(i)
            text_trans+=" "+ str(i) + ". " + texts[i]+ "| END OF TWEET."
            if i%7==0:
                all_text+=translate(text_trans, "en")
                text_trans=" "
                print("converting 7 text")
        
        all_text+=translate(text_trans,"en")
        #separte the strings
        trans=all_text.split("END OF TWEET")
        print("translating")
        print(len(ids))
        print(len(trans))
        f=0
        if len(ids)<= len(trans):
            f=len(ids)
        else:
            f=len(trans)
        print("Loss of "+ str(abs(len(ids)+1 -len(trans)))+ lang+ " Tweets")
        #update the tweets
        for tweet in final_tweets:
            for i in range(0,f):
                if tweet.id==ids[i]:
                    tweet.full_text=trans[i]
        
        return final_tweets
    
    def DownloadData(self, username, worksheet_list, start_date, end_date, list_variable, trans_tweets):
        #Twitter credentials for the app
        consumer_key = "XXX"
        consumer_secret = "XXX"
        access_key= "XXX"
        access_secret = "XXX"

        # Creating the authentication object
        auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
        # Setting your access token and secret
        auth.set_access_token(access_key, access_secret)
        # Creating the API object while passing in auth information
        api = tweepy.API(auth)
        
        #limit of tmpTweets is 20 tweets
        #use mode extended to get full tweet
        tweets = []
        print(username)
        f.write("\n"+username+"\n")
        tmpTweets = api.user_timeline(username, tweet_mode='extended')
        for tweet in tmpTweets:
            if tweet.created_at < end_date and tweet.created_at > start_date:
                tweets.append(tweet)
            
        limit= False
        while (tmpTweets[-1].created_at > start_date):
            print("Last Tweet @", tmpTweets[-1].created_at, " - fetching some more")
            tmpTweets = api.user_timeline(username, max_id = tmpTweets[-1].id, tweet_mode='extended')
            for tweet in tmpTweets:
                if tweet.created_at < end_date and tweet.created_at > start_date:
                    tweets.append(tweet)
                    if len(tweets) == 3200:
                        limit=True
                        print("Reached Twitter 3200 tweets limit time to change start date and end date, new end date is", tweet.created_at)
                        break
                if limit:
                    break
        
        f.write("Total number of Tweets including retweet without text= "+ str(len(tweets))+"\n")
        #removing duplicates
        final_tweets=[]
        h=0
        i=0
        j=0
        k=0
        l=0
        m=0
        retweet=0
        row=0
        hindi_tweet={'tweetid':[],'tweettext':[]}
        tamil_tweet={'tweetid':[],'tweettext':[]}
        gujrati_tweet={'tweetid':[],'tweettext':[]}
        marathi_tweet={'tweetid':[],'tweettext':[]}
        real_tweet_text=[]
        for tweet in tweets:
            retweet+=tweet.retweet_count
            if tweet not in final_tweets and tweet.lang=="en":
                final_tweets.append(tweet)
                real_tweet_text.append(tweet.full_text)
                i+=1
                tweet.full_text=tweet.full_text
            
            elif tweet not in final_tweets and tweet.lang=="hi":
                real_tweet_text.append(tweet.full_text)
                tweet.full_text = re.sub('https?://[A-Za -z0-9./]+','',tweet.full_text)
                hindi_tweet['tweetid'].append(tweet.id)
                hindi_tweet['tweettext'].append(tweet.full_text)
                final_tweets.append(tweet)
                j+=1
            
            
            elif tweet not in final_tweets and tweet.lang=="ta":
                real_tweet_text.append(tweet.full_text)
                tweet.full_text = re.sub('https?://[A-Za -z0-9./]+','',tweet.full_text)
                tamil_tweet['tweetid'].append(tweet.id)
                tamil_tweet['tweettext'].append(tweet.full_text)
                final_tweets.append(tweet)
                k+=1
                
            elif tweet not in final_tweets and tweet.lang=="gu":
                real_tweet_text.append(tweet.full_text)
                tweet.full_text = re.sub('https?://[A-Za -z0-9./]+','',tweet.full_text)
                gujrati_tweet['tweetid'].append(tweet.id)
                gujrati_tweet['tweettext'].append(tweet.full_text)
                #tweet.full_text=translate(tweet.full_text,'en')
                final_tweets.append(tweet)
                l+=1
            
            elif tweet not in final_tweets and tweet.lang=="mr":
                real_tweet_text.append(tweet.full_text)
                tweet.full_text = re.sub('https?://[A-Za -z0-9./]+','',tweet.full_text)
                #marathi_tweet['tweetid'].append(tweet.id)
                #marathi_tweet['tweettext'].append(tweet.full_text)
                tweet.full_text=translate(tweet.full_text,'en')
                final_tweets.append(tweet)
                m+=1
         
        if trans_tweets=="True":
            final_tweets=self.translator_text(final_tweets, hindi_tweet['tweetid'], hindi_tweet['tweettext'], " Hindi ")
            final_tweets=self.translator_text(final_tweets, tamil_tweet['tweetid'], tamil_tweet['tweettext'], " Tamil ")
            final_tweets=self.translator_text(final_tweets, gujrati_tweet['tweetid'], gujrati_tweet['tweettext']," Gujrati ")
            final_tweets=self.translator_text(final_tweets, marathi_tweet['tweetid'], marathi_tweet['tweettext']," Marathi ")

        f.write("Total number of retweets= "+ str(retweet)+ "\n")
        f.write("Number of English Tweets = " +str(i)+ "\n")
        f.write("Number of Hindi Tweets = "+ str(j)+"\n")
        f.write("Number of Tamil Tweets = " +str(k)+ "\n")
        f.write("Number of Gujrati Tweets = "+ str(l)+ "\n")
        f.write("Number of Marathi Tweets ="+ str(m)+ "\n")
        
        tweets = final_tweets
        #reversing so that we anaylse by plotting against given timeline
        tweets.reverse()
        real_tweet_text.reverse()
        for i in range(0,len(real_tweet_text)):
                worksheet_list[list_variable].write(i, 2, real_tweet_text[i])
        
        global length
        length=len(tweets)
        self.sentiment_analysis(tweets, username, worksheet_list, start_date, end_date, list_variable)

    def sentiment_analysis(self, tweets, username, worksheet_list, start_date, end_date, list_variable):
        # creating some variables to store info
        polarity = 0
        positive = 0
        wpositive = 0
        spositive = 0
        negative = 0
        wnegative = 0
        snegative = 0
        neutral = 0
    
        row=0
        retweet=0
        global user_dict
        # iterating through tweets fetched
        for tweet in tweets:
            worksheet_list[list_variable].write(row, 7, tweet.full_text)
            retweet+=tweet.retweet_count
            #remove @usernames and special characters(*,!,.,?)
            clean_text=' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t]) | (\w +:\ / \ / \S +)", " ", tweet.full_text).split())
            #remove urls and links
            clean_text = re.sub('https?://[A-Za -z0-9./]+','',clean_text)
            clean_text=clean_text.lower()
            #remove #xyz to xyz
            clean_text=re.sub("[^a-zA-Z]", " ", clean_text)
            #tokenize the words
            text_tokens = word_tokenize(clean_text)
            stopwords = nltk.corpus.stopwords.words('english')
            #as twitter api returns & as &amp
            new_stopwords=['amp', 'ji', 'shree','rs','shri', 'rt','mr', 'smt']
            for i in range(0,len(new_stopwords)):
                stopwords.append(new_stopwords[i])
            tokens_without_sw = [word for word in text_tokens if not word in stopwords]
            clean_text = (" ").join(tokens_without_sw)
            analysis = TextBlob(clean_text)
            # adding up polarities to find the average later

            polarity += analysis.sentiment.polarity
            
            
            user_dict[username].append(polarity)
            user_dict[username+'PolarityPerTweet'].append(analysis.sentiment.polarity)
            user_dict[username+'DataCreated'].append(tweet.created_at)
            user_dict[username+'average'].append(polarity/len(user_dict[username+'DataCreated']))
            if (analysis.sentiment.polarity == 0):
            # adding reaction of how people are reacting to find average later
                neutral += 1
                worksheet_list[list_variable].write(row, 6, "neutral")
            elif (analysis.sentiment.polarity > 0 and analysis.sentiment.polarity <= 0.3):
                wpositive += 1
                worksheet_list[list_variable].write(row, 6, "weak positive")
            elif (analysis.sentiment.polarity > 0.3 and analysis.sentiment.polarity <= 0.6):
                positive += 1
                worksheet_list[list_variable].write(row, 6, "positive")
            elif (analysis.sentiment.polarity > 0.6 and analysis.sentiment.polarity <= 1):
                spositive += 1
                worksheet_list[list_variable].write(row, 6, "strong positive")
            elif (analysis.sentiment.polarity > -0.3 and analysis.sentiment.polarity <= 0):
                wnegative += 1
                worksheet_list[list_variable].write(row, 6, "weak negative")
            elif (analysis.sentiment.polarity > -0.6 and analysis.sentiment.polarity <= -0.3):
                negative += 1
                worksheet_list[list_variable].write(row, 6, "negative")
            elif (analysis.sentiment.polarity > -1 and analysis.sentiment.polarity <= -0.6):
                snegative += 1
                worksheet_list[list_variable].write(row, 6, "strong negative")

            #writing data to workbook
            worksheet_list[list_variable].write_string(row, 0, str(tweet.id))
            worksheet_list[list_variable].write_string(row, 1, str(tweet.created_at))
            worksheet_list[list_variable].write(row, 3, clean_text)
            worksheet_list[list_variable].write_string(row, 4, str(tweet.in_reply_to_status_id))
            worksheet_list[list_variable].write_string(row, 5, str(analysis.sentiment.polarity))
            row += 1
                
        f.write("retweet of text tweets= "+str(retweet)+"\n")
        f.write("average retweest of text tweets="+ str(retweet/len(tweets))+ "\n")
        searchTerm= username
        NoOfTerms=len(tweets)
        # finding average of how people are reacting
        total_positive=positive+wpositive+spositive
        total_negative=negative+wnegative+snegative
        f.write("total_negative= "+ str(total_negative)+ "\n")
        f.write("total_positive= "+ str(total_positive)+ "\n")
        f.write("neutral="+ str(neutral)+ "\n")
        total_positive=self.percentage(total_positive,NoOfTerms)
        total_negative=self.percentage(total_negative, NoOfTerms)
        positive = self.percentage(positive, NoOfTerms)
        wpositive = self.percentage(wpositive, NoOfTerms)
        spositive = self.percentage(spositive, NoOfTerms)
        negative = self.percentage(negative, NoOfTerms)
        wnegative = self.percentage(wnegative, NoOfTerms)
        snegative = self.percentage(snegative, NoOfTerms)
        neutral = self.percentage(neutral, NoOfTerms)

        # finding average reaction
        polarity = polarity / NoOfTerms
        f.write("How people are reacting on " + searchTerm + " by analyzing " + str(NoOfTerms) + " tweets.\n")
        f.write("General Report: \n")

        if (polarity == 0):
            print("Neutral")
            f.write("Neutral")
        elif (polarity > 0 and polarity <= 0.3):
            print("Weakly Positive")
            f.write("Weakly Positive\n")
        elif (polarity > 0.3 and polarity <= 0.6):
            print("Positive")
            f.write("Positive\n")
        elif (polarity > 0.6 and polarity <= 1):
            print("Strongly Positive")
            f.write("Strongly Positive\n")
        elif (polarity > -0.3 and polarity <= 0):
            print("Weakly Negative")
            f.write("Weakly Negative\n")
        elif (polarity > -0.6 and polarity <= -0.3):
            print("Negative")
            f.write("Negative\n")
        elif (polarity > -1 and polarity <= -0.6):
            print("Strongly Negative")
            f.write("Strongly Negative\n")
        
        f.write("Detailed Report: \n")
        f.write(str(positive) + "% people thought it was positive\n")
        f.write(str(wpositive) + "% people thought it was weakly positive\n")
        f.write(str(spositive) + "% people thought it was strongly positive\n")
        f.write(str(negative) + "% people thought it was negative\n")
        f.write(str(wnegative) + "% people thought it was weakly negative\n")
        f.write(str(snegative) + "% people thought it was strongly negative\n")
        f.write(str(neutral) + "% people thought it was neutral\n")
        f.write(str(total_positive) + "% people thought it was overall positive\n")
        f.write(str(total_negative) + "% people thought it was overall negative\n")

        self.plotPieChart(positive, wpositive, spositive, negative, wnegative, snegative, neutral, searchTerm, NoOfTerms, username)
        
    # function to calculate percentage
    def percentage(self, part, whole):
        temp = 100 * float(part) / float(whole)
        return format(temp, '.2f')

    def plotPieChart(self, positive, wpositive, spositive, negative, wnegative, snegative, neutral, searchTerm, noOfSearchTerms, username):
        labels = ['Positive [' + str(positive) + '%]', 'Weakly Positive [' + str(wpositive) + '%]','Strongly Positive [' + str(spositive) + '%]', 'Neutral [' + str(neutral) + '%]',
                  'Negative [' + str(negative) + '%]', 'Weakly Negative [' + str(wnegative) + '%]', 'Strongly Negative [' + str(snegative) + '%]']
        sizes = [positive, wpositive, spositive, neutral, negative, wnegative, snegative]
        colors = ['yellowgreen','lightgreen','darkgreen', 'gold', 'red','lightsalmon','darkred']
        explode=[0.2,0.1,0.3,0,-0.2,-0.1,-0.3]
        patches, texts = plt.pie(sizes, colors=colors, labels=labels,startangle=90)
        plt.legend(patches, labels, loc="best")
        plt.title('Sentiment of ' + searchTerm + ' while tweeting ' + str(noOfSearchTerms) + ' Tweets.')
        plt.axis('equal')
        plt.tight_layout()
        plt.show(block=False)
        plt.savefig(username)
        plt.close()
    
            
def create_wordcloud(username, sheet):
        total_tokens=0
        comment_words = ''
        stopwords = set(STOPWORDS)
        
        for i in range(1, sheet.max_row+1):
          cell=sheet.cell(row=i, column=4)
          # typecaste each val to string
          val = str(cell.value)
          # split the value
          tokens = val.split()
          total_tokens+=len(tokens)
          comment_words += " ".join(tokens)+" "
          
        wordcloud = WordCloud(width = 800, height = 800,max_words=100,
                        background_color ='white',
                        stopwords = stopwords,
                        min_font_size = 10, collocations=False).generate(comment_words)
          
        # plot the WordCloud image
        f.write("Number of Total Words for "+str(username)+" "+ str(total_tokens) +"\n")
        plt.figure(figsize = (8, 8), facecolor = None)
        plt.imshow(wordcloud)
        plt.axis("off")
        plt.tight_layout(pad = 0)
        plt.savefig("wordcloud for "+ username)
        plt.close()


#gives per day data
def summation( data_created_at, polarity_per_tweet, start_date, next_date, end_date):
    polarity_per_day=0
    number_of_tweets=0
    for i in range(0,len(data_created_at)):
        if next_date <= end_date:
            if start_date < data_created_at[i] < next_date:
                polarity_per_day+=polarity_per_tweet[i]
                number_of_tweets+=1
    return polarity_per_day, number_of_tweets

#calculates per day data values
def day_wise_data(data_created_at, polarity_per_tweet, start_date, end_date):
    
    next_date=start_date+datetime.timedelta(days=1)
    tweet_per_day_polarity=[]
    tweets_per_day=[]
    polarity_per_day=0
    number_of_tweets=0
    dates=[]
    
    for i in range(0, len(data_created_at)):
        if next_date<=end_date:
                x,y=summation(data_created_at, polarity_per_tweet, start_date, next_date, end_date)
                if y!=0:
                    tweet_per_day_polarity.append(x/y)
                    tweets_per_day.append(y)
                    dates.append(start_date)
                else:
                    tweet_per_day_polarity.append(0)
                    tweets_per_day.append(0)
                    dates.append(start_date)
                start_date= next_date
                next_date= start_date +datetime.timedelta(days=1)
    
    total_polarity_sum=[]
    polarity_sum_avg=[]
    p=0
    z=0
    days=0
    for i in range(0, len(dates)):
        days+=1
        p+=(tweet_per_day_polarity[i]*tweets_per_day[i])
        z+=tweet_per_day_polarity[i]
        total_polarity_sum.append(p)
        polarity_sum_avg.append(z/days)
     
    return tweet_per_day_polarity, tweets_per_day, dates, total_polarity_sum, polarity_sum_avg
    

def plot_with_multiple_axes(username,x_value, y_value, x_value2, y_value2,x_value3, y_value3, xlabel, ylabel, xlabel2, ylabel2, linename, linename2,linename3, plotname, ymin , ymax):
    
        
    fig=plt.figure(figsize = (20, 10))
    ax=fig.add_subplot(111, label="1")
    ax2=fig.add_subplot(111, label="2", frame_on=False)
    x = [date2num(x) for x in x_value]
    ax.plot(x, y_value, color="C0", label=linename)
    ax.set_xlabel(xlabel, color="C0")
    ax.set_ylabel(ylabel, color="C0")
    #ax.axis(xmin=x_value[0], xmax=x_value[-1])
    ax.set_xticks(x[::7])
    ax.set_xticklabels([x.strftime("%d-%m") for x in x_value[::7]])
    if ymin!=0:
        ax.axis(ymin=ymin, ymax=ymax)
    ax.tick_params(axis='x', colors="C0")
    ax.tick_params(axis='y', colors="C0")
    ax.legend(loc=2)
    ax2.plot(x_value2, y_value2, color="C1", label=linename2)
    ax2.plot(x_value3, y_value3, color="C2", label=linename3)
    ax2.xaxis.tick_top()
    ax2.yaxis.tick_right()
    ax2.set_xticks([])
    #ax2.set_xlabel(xlabel2, color="C1")
    ax2.set_ylabel(ylabel2, color="C1")
    #ax2.axis(xmin=x_value[0], xmax=x_value[-1])
    ax2.xaxis.set_label_position('top')
    ax2.yaxis.set_label_position('right')
    ax2.tick_params(axis='x', colors="C1")
    ax2.tick_params(axis='y', colors="C1")
    ax2.legend(loc=1)
   
    plt.show(block=False)
    plt.savefig(plotname+username)
    plt.close()

def get_data_csv(list, date_start, date_end):
    start=0
    end=0
    for i in range(0, len(list)):
        if list[i]==date_start:
            start=i
        elif list[i]==date_end:
            end=i
    return start, end
    
def single_axis_plotter(xvalue, yvalue, yavg, username, date_list,names, xlabel, ylabel, fig_name, ymin, ymax):

    fig=plt.figure(figsize = (20, 10))
    ax=fig.add_subplot(111, label="1")
    color=['red','magenta','blue','green', 'yellow', 'cyan']
    x = [date2num(x) for x in xvalue[0]]
    for i in range(0, len(username)):
        ax.plot(x, yvalue[i], color=color[i], label=names[i])
        if i==0:
            ax.plot(xvalue[0], yavg, color='black', label='Cumulative Average' )
    color2=['yellow', 'white','cyan', 'yellow', 'white', 'cyan' ]
    for i in range(0,len(date_list)-1):
        s, e= get_data_csv(xvalue[0], date_list[i],date_list[i+1])
        ax.axvspan(xvalue[0][s],xvalue[0][e] , facecolor=color2[i], alpha=0.2)
    
    if ymin!=0:
        ax.axis(ymin=ymin, ymax=ymax)
    ax.set_xticks(x[::7])
    ax.set_xticklabels([x.strftime("%d-%m") for x in xvalue[0][::7]])
    ax.set_ylabel(ylabel, color="C0")
    ax.set_xlabel(xlabel, color="C0")
    ax.tick_params(axis='x', colors="C0")
    ax.tick_params(axis='y', colors="C0")
    ax.legend(loc='best')
    
    plt.show(block=False)
    plt.savefig(fig_name)
    plt.close()

def multiple_user_multiple_axes(xvalue, yvalue, xvalue2, yvalue2,y2linename, username, date_list, names, xlabel, ylabel, ylabel2, fig_name, ymin, ymax):
    fig=plt.figure(figsize = (20, 10))
    ax=fig.add_subplot(111, label="1")
    ax11=fig.add_subplot(111, label="1", frame_on=False)
    color=['red','magenta','blue','green', 'yellow', 'cyan']
    x = [date2num(x) for x in xvalue[0]]
    for i in range(0, len(username)):
        ax.plot(x, yvalue[i], color=color[i], label=names[i])
        if i==0:
            ax11.plot(xvalue2, yvalue2, color='black', label=y2linename)
    color2=['yellow', 'white','cyan', 'yellow', 'white', 'cyan' ]
    for i in range(0,len(date_list)-1):
        s, e= get_data_csv(xvalue[0], date_list[i],date_list[i+1])
        ax.axvspan(xvalue[0][s],xvalue[0][e] , facecolor=color2[i], alpha=0.2)
    
    if ymin!=0:
        ax.axis(ymin=ymin, ymax=ymax)
    ax.set_xticks(x[::7])
    ax.set_xticklabels([x.strftime("%d-%m") for x in xvalue[0][::7]])
    ax.set_ylabel(ylabel, color="C0")
    ax.set_xlabel(xlabel, color="C0")
    ax.tick_params(axis='x', colors="C0")
    ax.tick_params(axis='y', colors="C0")
    ax.legend(loc=2)
    
    ax11.xaxis.tick_top()
    ax11.yaxis.tick_right()
    ax11.set_xticks([])
    #ax11.set_xlabel("Days", color="C1")
    ax11.set_ylabel(ylabel2, color="C1")
    ax11.xaxis.set_label_position('top')
    ax11.yaxis.set_label_position('right')
    ax11.tick_params(axis='x', colors="C1")
    ax11.tick_params(axis='y', colors="C1")
    ax11.legend(loc=1)
    
    plt.show(block=False)
    plt.savefig(fig_name)
    plt.close()
    
def average_comparison_plot(user_dict, start_date, end_date, state_comp, individual_comp, date_list):
    names=[]
    day_polarity=[]
    tweet_day=[]
    dates_all=[]
    sum_pol=[]
    avg=[]
    color=['red','magenta','blue','green', 'yellow', 'cyan']
    xticks=[]
    average_all=[]
    sum_all=[]
    day_polarity_all=[]
    tweet_all=[]
    
    for i in range(0,len(username)):
        names.append(username[i])
        tweet_per_day_polarity, tweets_per_day, dates, total_polarity_sum, polarity_sum_avg = day_wise_data(user_dict[username[i]+'DataCreated'], user_dict[username[i]+'PolarityPerTweet'], start_date, end_date)
        day_polarity.append(tweet_per_day_polarity)
        tweet_day.append(tweets_per_day)
        dates_all.append(dates)
        sum_pol.append(total_polarity_sum)
        avg.append(polarity_sum_avg)
        worksheet1.write_string(0,i+3,str(username[i]))
        worksheet1.write_column(1,i+3, polarity_sum_avg)
        
    #sum of all polarites of all leaders
    tweet_all=[sum(x) for x in zip(*tweet_day)]
    average_all=[sum(x) for x in zip(*avg)]
    sum_all=[sum(x) for x in zip(*sum_pol)]
    day_polarity_all=[sum(x) for x in zip(*day_polarity)]
  
    for i in range(0, len(sum_all)):
        average_all[i]=average_all[i]/len(username)
        sum_all[i]=sum_all[i]/len(username)
        day_polarity_all[i]=day_polarity_all[i]/len(username)
        tweet_all[i]=tweet_all[i]/len(username)
        
    worksheet1.write_column(1, 0, sum_all)
    worksheet1.write_column(1, 1, average_all)
    worksheet1.write_column(1, 2, day_polarity_all)
    worksheet1.write_string(0,0,str("Sum of all polarity"))
    worksheet1.write_string(0,1,str("Average of all polarity"))
    worksheet1.write_string(0,2,str("Polarity per day for all"))
    
    #single_axis_plotter(xvalue, yvalue, yavg, username, date_list, xlabel, ylabel, fig_name, ymin , ymax)
    single_axis_plotter(dates_all, avg, average_all, username, date_list, names ,"Days", "Average of Tweet Polarity", "Average Plot", -1.1, 1.1)
    single_axis_plotter(dates_all, sum_pol, sum_all, username, date_list,names, "Days", "Sum of polarities", "Sum Plot",0,0)
    single_axis_plotter(dates_all, day_polarity, day_polarity_all, username, date_list,names, "Time", "Polarity per day", " Day Polarity Plot",-1.1,1.1)
    single_axis_plotter(dates_all, tweet_day, tweet_all, username, date_list,names, "Days", "Number of Tweets", "Number of Tweets Plot",0,0)

    #data for national corona plots
    file=pd.read_csv('nation_level_daily.csv', header=0)
    s_date=start_date.strftime('%d-%b')
    e_date=end_date.strftime('%d-%b')
    corona_cases=[]
    total_deceased=[]
    daily_deceased=[]
    #getting start and end dates
    p,q=get_data_csv(file.Date, s_date, e_date)
    #adding values to the list of cases
    for i in range(p, q):
        corona_cases.append(file.totalconfirmed[i])
        total_deceased.append(file.totaldeaths[i])
        daily_deceased.append(file.dailydeceased[i])
    
    #multiple_user_multiple_axes(xvalue, yvalue, xvalue2, yvalue2, y2linename,username, date_list,names, xlabel, ylabel, ylabel2, fig_name, ymin, ymax)
    #username, datelist and names are code variables
    multiple_user_multiple_axes(dates_all, day_polarity, dates_all[0], daily_deceased, "Daily Deaths",username, date_list,names, "Days", "Polarity per day", "Number of cases", "Death vs Polarity per day Plot", -1.1, 1.1)
    multiple_user_multiple_axes(dates_all, avg, dates_all[0], daily_deceased, "Daily Deaths",username, date_list,names, "Days", "Average per day", "Number of cases", "Daily Deaths vs Average Plot", -1.1, 1.1)
    
    
    #plot_with_multiple_axes(username,x_value, y_value, x_value2, y_value2,x_value3, y_value3, xlabel, ylabel, xlabel2, ylabel2, linename, linename2,linename3, plotname, ymin , ymax):
    if individual_comp=="True":
        for i in range(0, len(username)):
               plot_with_multiple_axes(username[i], dates_all[i], day_polarity[i], dates_all[i], daily_deceased,dates_all[i], total_deceased,"days","polarity per day",  "days","Number of cases", username[i], "Daily Deaths","Total Deaths", "Polarity per day vs daily_deceased for ", -1.1, 1.1 )
               plot_with_multiple_axes(username[i], dates_all[i], sum_pol[i], dates_all[i], daily_deceased, dates_all[i], total_deceased, "days","sum of polarity of tweets","days", "Polarity of tweets per day",  username[i],"Daily Deaths","Total Deaths", "Polarity sum vs daily_deceased for ", 0, 0 )
               plot_with_multiple_axes(username[i], dates_all[i], avg[i], dates_all[i], daily_deceased,dates_all[i], total_deceased,  "days","Average polarity","days", "Polarity of tweets per day",  username[i], "Daily Deaths","Total Deaths", "Average Polarity vs daily_deceased for ", -1.1, 1.1 )
               
               
               
# get the state wise data
def get_state_data(state_name):
    state=pd.read_csv('state.csv', index_col="State_Name")
    state_cases=state.loc[state_name]
    df=pd.DataFrame(state_cases)
    df['Date']=pd.to_datetime(df['Date'])
    df=df.sort_values(by='Date')
    state_dict={'dates':[],'deaths':[],'confirmed':[]}
    state_dict['confirmed'].append(df['Confirmed'].tolist())
    state_dict['dates'].append(df['Date'].tolist())
    state_dict['deaths'].append(df['Deceased'].tolist())
    dates=[]
    state_corona_cases_daily=[]
    state_daily_deceased=[]
    ds,de=get_data_csv(state_dict['dates'][0], start_date,end_date)
    for i in range(ds,de):
        dates.append(state_dict['dates'][0][i])
        state_corona_cases_daily.append(state_dict['confirmed'][0][i])
        state_daily_deceased.append(state_dict['deaths'][0][i])

    return dates, state_corona_cases_daily, state_daily_deceased

#plotting state data
def state_plots(user_dict, state_list, start_date, end_date):
    names=[]
    day_polarity=[]
    tweet_day=[]
    dates_all=[]
    sum=[]
    avg=[]
    color=['red','cyan','blue','green', 'yellow', 'magenta']
    xticks=[]
    for i in range(0,len(username)):
        names.append(username[i])
        tweet_per_day_polarity, tweets_per_day, dates, total_polarity_sum, polarity_sum_avg = day_wise_data(user_dict[username[i]+'DataCreated'], user_dict[username[i]+'PolarityPerTweet'], start_date, end_date)
    
        day_polarity.append(tweet_per_day_polarity)
        tweet_day.append(tweets_per_day)
        dates_all.append(dates)
        sum.append(total_polarity_sum)
        avg.append(polarity_sum_avg)

    state_dates=[]
    state_corona_cases_daily=[]
    state_daily_deceased=[]
    for i in range(len(state_list)):
        dates, cases, deaths=get_state_data(state_list[i])
        state_dates.append(dates)
        state_corona_cases_daily.append(cases)
        state_daily_deceased.append(deaths)
    
    #plot_with_multiple_axes(username,x_value, y_value, x_value2, y_value2,x_value3, y_value3, xlabel, ylabel, xlabel2, ylabel2, linename, linename2,linename3, plotname, ymin , ymax):
    for i in range(0, len(username)):
        plot_with_multiple_axes(state_list[i], dates_all[i], day_polarity[i], dates_all[i], state_corona_cases_daily[i],dates_all[i], state_daily_deceased[i],"days","polarity per day",  "days","Cases", username[i], "Daily Cases","Daily Deaths", "Polarity vs state_daily_cases for ", -1.1, 1.1 )
        plot_with_multiple_axes(state_list[i], dates_all[i], sum[i], dates_all[i], state_corona_cases_daily[i], dates_all[i], state_daily_deceased[i], "days","polarity sum","days", "Cases",  username[i], "Daily Cases","Daily Deaths", "Sum vs state_daily_deceased for ", 0, 0 )
        plot_with_multiple_axes(state_list[i], dates_all[i], avg[i], dates_all[i], state_corona_cases_daily[i], dates_all[i], state_daily_deceased[i], "days","Average polarity","days", "Cases",  username[i], "Daily Cases","Daily Deaths", "Average vs state_daily_deceased for ", -1.1, 1.1 )
        

def ObtainDate():
    isValid=False
    while not isValid:
        userIn =input("Type Date dd/mm/yy: ")
        try: # strptime throws an exception if the input doesn't match the pattern
            d = datetime.datetime.strptime(userIn, "%d/%m/%y")
            isValid=True
        except:
            print("Oops, try again!\n")
    return d
    
if __name__== "__main__":
   print("Enter Start Date After 15/3/20: ")
   start_date=ObtainDate()
   print("Start Date ",start_date)
   print("Enter End Date Before 1/7/20:")
   #print("If you want data till 31/5/20 then enter the next date that is 1/6/20")
   end_date=ObtainDate()
   print("Enter the number of users you want to analyse :")
   users=int(input())
   username=[]
   for i in range(0, users):
       print("Enter Twitter Ids of user ", i+1)
       username.append(str(input()))
   print(username)
   trans_tweets=input("Do you want to translate tweets to English (True/False): ")
   individual_comp=input("Do you want to compare individual accounts with corona cases (True/False): ")
   state_comp=input("Do you want to compare state wise data(True/False): ")
   if state_comp=="True":
       print("Format for state name: andhra pradesh as Andhra Pradesh and delhi as Delhi" )
       state_list=[]
       for i in range(0, len(username)):
           print("Enter State for "+ username[i])
           state_list.append(str(input()))
       print(state_list)
   date_list=[]
   date_list.append(start_date)
   number = input("Number of periods you want to highlight :")
   print("Enter dates between "+ str(start_date.strftime('%d/%m/%y'))+" and "+str(end_date.strftime('%d/%m/%y')))
   mid_date=start_date
   end_date=end_date+datetime.timedelta(days=1)
   for i in range(1, int(number)+1):
       print("Enter date number"+ str(i)+ " after "+ str(mid_date.strftime('%d/%m/%y')))
       date_list.append(ObtainDate())
       mid_date=date_list[i]
   f=open("terminal_output.txt","w")
   print(date_list)
   workbook=xlsxwriter.Workbook("Tweets.xlsx")
   worksheet_list=[]
   list_variable=0
   for i in range(len(username)):
       worksheet=workbook.add_worksheet(username[i])
       worksheet_list.append(worksheet)
   
   worksheet1=workbook.add_worksheet('Regression data')
   length=0
   user_dict={}
   for i in range(0,len(username)):
       user_dict.update({username[i]:[], username[i]+'PolarityPerTweet':[], username[i]+'DataCreated':[], username[i]+'average':[]})
      
   time_=[]
   polarity_=[]
   sentiment_=[]
   sa = SentimentAnalysis()
   for i in range(0, len(username)):
       sa.DownloadData(username[i], worksheet_list, start_date, end_date,i, trans_tweets)
   
print("Excel file ready")
print("Plotting Graphs")
if state_comp=="True":
    state_plots(user_dict, state_list, start_date, end_date)
average_comparison_plot(user_dict, start_date, end_date, state_comp, individual_comp, date_list)
print("Complete")
workbook.close()
print("Creating Wordcloud")
filename=str('Tweets.xlsx')
df = openpyxl.load_workbook(filename)
for i in range(0, len(username)):
    wb=df[username[i]]
    create_wordcloud(username[i], wb)
f.close()

    
